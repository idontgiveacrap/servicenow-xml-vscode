#!/usr/bin/env python3
"""
Convert the ServiceNow scripting reference workbook into MCP-ready JSON.gz.

Source workbook is human-oriented (cover sheet, title rows, multi-sheet layout).
This script strips presentation rows, normalizes column names, and writes a single
UTF-8 JSON document gzipped for the servicenow-xml-scripting MCP server.

Usage:
  python scripts/pack-scripting-reference.py [path/to/workbook.xlsx]
  python scripts/pack-scripting-reference.py --out cursor-plugins/servicenow-xml/data/scripting_reference.json.gz
"""

from __future__ import annotations

import argparse
import gzip
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: python -m pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = (
    REPO_ROOT
    / "cursor-plugins"
    / "servicenow-xml"
    / "data"
    / "scripting_reference.json.gz"
)

# Common mojibake / replacement-char patterns from the workbook export.
_REPLACEMENTS = (
    ("\ufffd", "—"),
    ("No — ", "No — "),
    ("Mixed — ", "Mixed — "),
)


def _cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    for old, new in _REPLACEMENTS:
        text = text.replace(old, new)
    # Collapse accidental double dashes introduced by prior replacements.
    text = text.replace("——", "—")
    return text


def _snake(header: str) -> str:
    text = _cell(header).lower()
    text = text.replace("?", "")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _row_values(row: tuple[Any, ...], width: int) -> list[str]:
    vals = [_cell(c) for c in row]
    if len(vals) < width:
        vals.extend([""] * (width - len(vals)))
    return vals[:width]


def _find_header_row(
    rows: list[tuple[Any, ...]], first_header: str
) -> tuple[int, list[str]] | None:
    want = first_header.strip().lower()
    for i, row in enumerate(rows):
        if not row:
            continue
        first = _cell(row[0]).lower()
        if first == want:
            headers = [_snake(_cell(c)) for c in row if _cell(c)]
            return i, headers
    return None


def _table_after_header(
    rows: list[tuple[Any, ...]],
    first_header: str,
    *,
    stop_at_blank: bool = False,
) -> list[dict[str, str]]:
    found = _find_header_row(rows, first_header)
    if not found:
        raise ValueError(f"Header starting with {first_header!r} not found")
    header_idx, headers = found
    width = len(headers)
    out: list[dict[str, str]] = []
    for row in rows[header_idx + 1 :]:
        vals = _row_values(row, width)
        if not any(vals):
            if stop_at_blank and out:
                break
            continue
        item = {headers[i]: vals[i] for i in range(width)}
        out.append(item)
    return out


def _section_after_label(
    rows: list[tuple[Any, ...]], label: str, first_header: str
) -> list[dict[str, str]]:
    """Parse a secondary table that starts after a section label row."""
    label_n = label.strip().lower()
    start = None
    for i, row in enumerate(rows):
        if _cell(row[0] if row else "").lower() == label_n:
            start = i
            break
    if start is None:
        return []
    return _table_after_header(rows[start:], first_header, stop_at_blank=True)


def _cover_meta(rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    title = ""
    subtitle = ""
    warnings: list[str] = []
    for i, row in enumerate(rows):
        first = _cell(row[0]) if row else ""
        if i == 0 and first:
            title = first
        elif i == 1 and first:
            subtitle = first
        elif first.upper() == "READ THIS FIRST":
            nxt = rows[i + 1] if i + 1 < len(rows) else ()
            warning = _cell(nxt[0]) if nxt else ""
            if warning:
                warnings.append(warning)
    return {
        "title": title or "ServiceNow Runtime Reference",
        "subtitle": subtitle,
        "warnings": warnings,
    }


def _runtime_catalog_story(rows: list[tuple[Any, ...]]) -> str:
    for i, row in enumerate(rows):
        if _cell(row[0]).upper() == "STORY IN BRIEF":
            nxt = rows[i + 1] if i + 1 < len(rows) else ()
            return _cell(nxt[0]) if nxt else ""
    return ""


def _rename_keys(rows: list[dict[str, str]], mapping: dict[str, str]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    for row in rows:
        item: dict[str, str] = {}
        for key, value in row.items():
            item[mapping.get(key, key)] = value
        out.append(item)
    return out


def pack_workbook(xlsx_path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    def sheet_rows(name: str) -> list[tuple[Any, ...]]:
        if name not in wb.sheetnames:
            raise ValueError(f"Missing sheet {name!r}; found {wb.sheetnames}")
        return list(wb[name].iter_rows(values_only=True))

    cover = _cover_meta(sheet_rows("Cover"))
    catalog_rows = sheet_rows("Runtime Catalog")
    runtime_catalog = _table_after_header(
        catalog_rows, "Pattern / family", stop_at_blank=True
    )
    runtime_catalog = _rename_keys(
        runtime_catalog,
        {
            "pattern_family": "pattern",
            "what_it_represents": "represents",
            "examples_from_export": "examples",
            "supported_contract": "supported_contract",
        },
    )
    documentation_workflow = _section_after_label(
        catalog_rows, "SAFE DOCUMENTATION WORKFLOW", "Step"
    )
    reference_sources: list[dict[str, str]] = []
    capturing_sources = False
    for row in catalog_rows:
        first = _cell(row[0] if row else "")
        if first.upper() == "REFERENCE SOURCES":
            capturing_sources = True
            continue
        if not capturing_sources:
            continue
        vals = _row_values(row, 3)
        if not any(vals):
            if reference_sources:
                break
            continue
        reference_sources.append({"name": vals[0], "note": vals[1], "url": vals[2]})

    runtime_items = _table_after_header(
        sheet_rows("Runtime Items"), "Name", stop_at_blank=True
    )
    runtime_items = _rename_keys(
        runtime_items,
        {"supported_contract": "supported_contract"},
    )

    server = _table_after_header(sheet_rows("server"), "Property", stop_at_blank=True)
    server = _rename_keys(
        server,
        {
            "property": "name",
            "documentation_link": "documentation_url",
        },
    )

    useful = _table_after_header(
        sheet_rows("Useful scripts"), "script", stop_at_blank=True
    )
    undocumented = _table_after_header(
        sheet_rows("Misc undocumented"), "API", stop_at_blank=True
    )
    undocumented = _rename_keys(
        undocumented,
        {
            "api": "api",
            "functions_methods_endpoints": "members",
        },
    )
    ui_builder = _table_after_header(
        sheet_rows("UI Builder-workspace"), "script", stop_at_blank=True
    )

    wb.close()

    return {
        "version": 1,
        "format": "servicenow_scripting_reference_json",
        "source_workbook": xlsx_path.name,
        "title": cover["title"],
        "subtitle": cover["subtitle"],
        "warnings": cover["warnings"],
        "documentation_workflow": documentation_workflow,
        "reference_sources": reference_sources,
        "runtime_catalog": {
            "story": _runtime_catalog_story(catalog_rows),
            "families": runtime_catalog,
        },
        "runtime_items": runtime_items,
        "server": server,
        "useful_scripts": useful,
        "undocumented": undocumented,
        "ui_builder": ui_builder,
        "counts": {
            "runtime_catalog_families": len(runtime_catalog),
            "runtime_items": len(runtime_items),
            "server": len(server),
            "useful_scripts": len(useful),
            "undocumented": len(undocumented),
            "ui_builder": len(ui_builder),
            "documentation_workflow_steps": len(documentation_workflow),
            "reference_sources": len(reference_sources),
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "workbook",
        nargs="?",
        default="",
        help="Path to ServiceNow scripting reference.xlsx",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=DEFAULT_OUT,
        help=f"Output .json.gz path (default: {DEFAULT_OUT})",
    )
    args = parser.parse_args()

    workbook = Path(args.workbook) if args.workbook else None
    if workbook is None or not workbook.is_file():
        print(
            "Workbook path required and must exist.\n"
            "Example:\n"
            '  python scripts/pack-scripting-reference.py '
            '"C:/path/ServiceNow scripting reference.xlsx"',
            file=sys.stderr,
        )
        return 2

    data = pack_workbook(workbook)
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    args.out.parent.mkdir(parents=True, exist_ok=True)
    with gzip.open(args.out, "wb") as handle:
        handle.write(payload)

    counts = data["counts"]
    print(f"Wrote {args.out} ({len(payload)} bytes uncompressed)")
    for key, value in counts.items():
        print(f"  {key}: {value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
